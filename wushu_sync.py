"""
2025 KL Wushu Championship — Live Results Sync
================================================
Logs into cmf-wushu.com, scrapes all 148 ranking pages,
writes to Excel + Google Sheets every 5 minutes.

SETUP (run once):
  pip install requests beautifulsoup4 gspread google-auth openpyxl schedule

RUN:
  python wushu_sync.py

Leave it running during the event. Press Ctrl+C to stop.
"""

import time
import schedule
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import json
import subprocess
from config import USERNAME, PASSWORD

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

BASE_URL   = "https://cmf-wushu.com/robin2026"
LOGIN_URL  = f"{BASE_URL}/login_admin.php"

EXCEL_FILE = "wushu_results.xlsx"
JSON_FILE  = "results_data.json"

# Google Sheets (optional) — leave GOOGLE_SHEET_ID as None to skip
GOOGLE_CREDS_FILE = "credentials.json"
GOOGLE_SHEET_ID   = None   # e.g. "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgVE2upms"

INTERVAL_MINUTES = 5

# ─────────────────────────────────────────────────────────────────────────────
# ALL 392 CATEGORY IDs  (from Claude Code scan, grouped by age division)
# ─────────────────────────────────────────────────────────────────────────────

CATEGORIES = {
    "U6": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20],
    "U8": [21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40,
           81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100],
    "U10": [41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60,
            101, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111, 112, 113, 114, 115, 116, 117, 118, 119, 120],
    "U12": [61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80,
            121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140,
            161, 162, 163, 164, 165, 166, 167, 168, 169, 170, 171, 172, 173, 174, 175, 176, 177, 178, 179, 180],
    "U15": [141, 142, 143, 144, 145, 146, 147, 148, 149, 150, 151, 152, 153, 154, 155, 156, 157, 158, 159, 160,
            181, 182, 183, 184, 185, 186, 187, 188, 189, 190, 191, 192, 193, 194, 195, 196, 197, 198, 199, 200,
            241, 242, 243, 244, 245, 246, 247, 248, 249, 250, 251, 252, 253, 254, 255, 256, 257, 258, 259, 260,
            261, 262, 307, 308, 309, 310, 311, 312, 313, 314, 315, 316, 317, 318, 319, 320, 321, 322, 323, 324,
            325, 326, 327, 328],
    "U18": [201, 202, 203, 204, 205, 206, 207, 208, 209, 210, 211, 212, 213, 214, 215, 216, 217, 218, 219, 220,
            263, 264, 265, 266, 267, 268, 269, 270, 271, 272, 273, 274, 275, 276, 277, 278, 279, 280, 281, 282,
            283, 284, 329, 330, 331, 332, 333, 334, 335, 336, 337, 338, 339, 340, 341, 342, 343, 344, 345, 346,
            347, 348, 349, 350, 373, 374, 375, 376, 377, 378, 379, 380, 381, 382, 383, 384, 385, 386, 387, 388,
            389, 390, 391, 392],
    "U25": [221, 222, 223, 224, 225, 226, 227, 228, 229, 230, 231, 232, 233, 234, 235, 236, 237, 238, 239, 240,
            285, 286, 287, 288, 289, 290, 291, 292, 293, 294, 295, 296, 297, 298, 299, 300, 301, 302, 303, 304,
            305, 306, 351, 352, 353, 354, 355, 356, 357, 358, 359, 360, 361, 362, 363, 364, 365, 366, 367, 368,
            369, 370, 371, 372],
}

# Flat list of all IDs (for scanning)
ALL_IDS = [id_ for ids in CATEGORIES.values() for id_ in ids]

# Reverse lookup: id → section name
ID_TO_SECTION = {id_: sec for sec, ids in CATEGORIES.items() for id_ in ids}

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────

def create_session():
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })
    login_page = session.get(LOGIN_URL, timeout=15)
    soup = BeautifulSoup(login_page.text, "html.parser")

    payload = {}
    for inp in soup.find_all("input"):
        name = inp.get("name")
        if name:
            payload[name] = inp.get("value", "")
    for key in list(payload.keys()):
        kl = key.lower()
        if any(x in kl for x in ["user", "login", "email"]):
            payload[key] = USERNAME
        if any(x in kl for x in ["pass", "pwd"]):
            payload[key] = PASSWORD
    payload.setdefault("username", USERNAME)
    payload.setdefault("password", PASSWORD)

    resp = session.post(LOGIN_URL, data=payload, timeout=15, allow_redirects=True)
    if "login" in resp.url.lower() and "login" in resp.text.lower()[:300]:
        print("  ⚠  Login may have failed — check credentials")
    else:
        print("  ✓  Logged in")
    return session

# ─────────────────────────────────────────────────────────────────────────────
# SCRAPE ONE CATEGORY
# ─────────────────────────────────────────────────────────────────────────────

def scrape_category(session, cat_id):
    url  = f"{BASE_URL}/show_ranking.php?id={cat_id}"
    resp = session.get(url, timeout=15, allow_redirects=True)
    if "login" in resp.url.lower():
        return None, None   # session expired

    soup = BeautifulSoup(resp.text, "html.parser")

    # Page title / category name
    name = None
    for tag in ["h1", "h2", "h3", "title"]:
        el = soup.find(tag)
        if el:
            text = el.get_text(strip=True)
            if text and len(text) > 2:
                name = text
                break
    name = name or f"Category {cat_id}"

    table = soup.find("table")
    if not table:
        return name, []

    rows    = table.find_all("tr")
    headers = []
    results = []

    for i, row in enumerate(rows):
        cells  = row.find_all(["th", "td"])
        values = [c.get_text(strip=True) for c in cells]
        if not values:
            continue
        if i == 0 or all(c.name == "th" for c in cells):
            headers = values
        else:
            if any(v for v in values):
                entry = {headers[j] if j < len(headers) else f"col_{j}": v
                         for j, v in enumerate(values)}
                results.append(entry)

    return name, results

# ─────────────────────────────────────────────────────────────────────────────
# WRITE EXCEL
# ─────────────────────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
SEC_FILL   = PatternFill("solid", fgColor="2E75B6")
SEC_FONT   = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")
GOLD_FILL  = PatternFill("solid", fgColor="FFD700")
SIL_FILL   = PatternFill("solid", fgColor="C0C0C0")
BRZ_FILL   = PatternFill("solid", fgColor="CD7F32")

def medal_fill(rank_str):
    r = str(rank_str).strip()
    if r in ("1", "1st"):  return GOLD_FILL
    if r in ("2", "2nd"):  return SIL_FILL
    if r in ("3", "3rd"):  return BRZ_FILL
    return None

def write_excel(all_data, now_str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.append(["Section", "ID", "Category", "Competitors", "Last Updated"])
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20

    current_section = None
    row_num = 2
    for section, cat_id, cat_name, results in all_data:
        if section != current_section:
            ws.cell(row=row_num, column=1, value=section)
            for c in range(1, 6):
                ws.cell(row=row_num, column=c).fill = SEC_FILL
                ws.cell(row=row_num, column=c).font = SEC_FONT
            row_num += 1
            current_section = section
        ws.cell(row=row_num, column=1, value=section)
        ws.cell(row=row_num, column=2, value=cat_id)
        ws.cell(row=row_num, column=3, value=cat_name)
        ws.cell(row=row_num, column=4, value=len(results))
        ws.cell(row=row_num, column=5, value=now_str)
        if len(results) == 0:
            for c in range(1, 6):
                ws.cell(row=row_num, column=c).font = Font(color="999999", italic=True)
        row_num += 1

    # ── One sheet per section ────────────────────────────────────────────────
    section_sheets = {}
    for section, cat_id, cat_name, results in all_data:
        sheet_name = section[:31]
        if sheet_name not in section_sheets:
            ws2 = wb.create_sheet(sheet_name)
            ws2.append([f"Section: {section}", "", "", "", f"Updated: {now_str}"])
            ws2.merge_cells("A1:D1")
            for c in range(1, 6):
                ws2.cell(1, c).fill = HDR_FILL
                ws2.cell(1, c).font = HDR_FONT
            section_sheets[sheet_name] = {"ws": ws2, "row": 2}

        ws2   = section_sheets[sheet_name]["ws"]
        start = section_sheets[sheet_name]["row"]

        # Category sub-header
        ws2.cell(start, 1, value=f"[{cat_id}] {cat_name}")
        ws2.cell(start, 1).fill = PatternFill("solid", fgColor="BDD7EE")
        ws2.cell(start, 1).font = Font(bold=True, size=10)
        ws2.merge_cells(f"A{start}:F{start}")
        start += 1

        if not results:
            ws2.cell(start, 1, value="— No results yet —")
            ws2.cell(start, 1).font = Font(italic=True, color="999999")
            section_sheets[sheet_name]["row"] = start + 2
            continue

        headers = list(results[0].keys())
        for j, h in enumerate(headers, 1):
            cell = ws2.cell(start, j, value=h)
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center")
        start += 1

        for i, row in enumerate(results):
            for j, h in enumerate(headers, 1):
                ws2.cell(start, j, value=row.get(h, ""))
            rank_val = list(row.values())[0] if row else ""
            fill = medal_fill(rank_val) or (ALT_FILL if i % 2 == 0 else None)
            if fill:
                for j in range(1, len(headers) + 1):
                    ws2.cell(start, j).fill = fill
            start += 1

        section_sheets[sheet_name]["row"] = start + 1

    # Auto-fit columns on each sheet
    for ws3 in wb.worksheets:
        for col in ws3.columns:
            try:
                col_letter = col[0].column_letter
            except AttributeError:
                continue  # skip merged cells
            max_len = max((len(str(c.value or "")) for c in col if hasattr(c, "value")), default=8)
            ws3.column_dimensions[col_letter].width = min(max_len + 4, 45)

    wb.save(EXCEL_FILE)
    print(f"  ✓  Excel saved → {EXCEL_FILE}")

# ─────────────────────────────────────────────────────────────────────────────
# WRITE JSON
# ─────────────────────────────────────────────────────────────────────────────

def write_json(all_data, now_str):
    payload = {
        "updated": now_str,
        "sections": [
            {"section": section, "cat_id": cat_id, "cat_name": cat_name, "results": results}
            for section, cat_id, cat_name, results in all_data
        ],
    }
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  ✓  JSON saved → {JSON_FILE}")

# ─────────────────────────────────────────────────────────────────────────────
# GIT AUTO-COMMIT + PUSH
# ─────────────────────────────────────────────────────────────────────────────

def git_commit_and_push(now_str):
    try:
        subprocess.run(["git", "add", JSON_FILE], check=True)
        diff = subprocess.run(["git", "diff", "--cached", "--quiet"])
        if diff.returncode == 0:
            print("  →  No changes to commit")
            return
        subprocess.run(["git", "commit", "-m", f"Sync results {now_str}"], check=True)
        subprocess.run(["git", "push"], check=True)
        print("  ✓  Committed and pushed to GitHub")
    except subprocess.CalledProcessError as e:
        print(f"  ✗  Git commit/push failed: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# WRITE GOOGLE SHEETS (optional)
# ─────────────────────────────────────────────────────────────────────────────

def write_google_sheets(all_data, now_str):
    if not GOOGLE_SHEET_ID:
        return
    try:
        creds = Credentials.from_service_account_file(
            GOOGLE_CREDS_FILE,
            scopes=["https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive"]
        )
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(GOOGLE_SHEET_ID)
    except Exception as e:
        print(f"  ⚠  Google Sheets auth failed: {e}")
        return

    # Group by section → one tab per section
    sections = {}
    for section, cat_id, cat_name, results in all_data:
        sections.setdefault(section, []).append((cat_id, cat_name, results))

    for section, cats in sections.items():
        try:
            tab_name = section[:100]
            try:
                ws = sh.worksheet(tab_name)
                ws.clear()
            except gspread.WorksheetNotFound:
                ws = sh.add_worksheet(title=tab_name, rows=500, cols=20)

            out = [[f"Section: {section}", f"Updated: {now_str}"], []]
            for cat_id, cat_name, results in cats:
                out.append([f"[{cat_id}] {cat_name}"])
                if not results:
                    out.append(["— No results yet —"])
                else:
                    out.append(list(results[0].keys()))
                    for row in results:
                        out.append(list(row.values()))
                out.append([])

            ws.update("A1", out)
            print(f"  ✓  Sheets updated: {section}")
        except Exception as e:
            print(f"  ⚠  Sheets error ({section}): {e}")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN SYNC
# ─────────────────────────────────────────────────────────────────────────────

def sync():
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n[{now_str}] Syncing {len(ALL_IDS)} categories...")

    try:
        session = create_session()
    except Exception as e:
        print(f"  ✗  Login failed: {e}")
        return

    all_data = []   # list of (section, cat_id, cat_name, results)
    expired  = False

    for section, ids in CATEGORIES.items():
        for cat_id in ids:
            try:
                cat_name, results = scrape_category(session, cat_id)
                if cat_name is None:   # session expired mid-run
                    if not expired:
                        print(f"  ⚠  Session expired at ID {cat_id}, re-logging in...")
                        session = create_session()
                        expired = True
                        cat_name, results = scrape_category(session, cat_id)
                if results is not None:
                    all_data.append((section, cat_id, cat_name, results))
                    total = len(results)
                    print(f"  ✓  [{cat_id:>3}] {cat_name[:45]:<45} {total} results")
            except Exception as e:
                print(f"  ✗  [{cat_id}] Error: {e}")
            time.sleep(0.2)   # polite delay

    # Write outputs
    if all_data:
        try:
            write_excel(all_data, now_str)
        except Exception as e:
            print(f"  ✗  Excel error: {e}")
        try:
            write_google_sheets(all_data, now_str)
        except Exception as e:
            print(f"  ✗  Google Sheets error: {e}")
        try:
            write_json(all_data, now_str)
            git_commit_and_push(now_str)
        except Exception as e:
            print(f"  ✗  JSON/git error: {e}")

    print(f"  → Done. Next sync in {INTERVAL_MINUTES} min.")

# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("  2025 KL Wushu Championship — Live Results Sync")
    print("=" * 60)
    print(f"  Categories : {len(ALL_IDS)} across {len(CATEGORIES)} sections")
    print(f"  Interval   : every {INTERVAL_MINUTES} minutes")
    print(f"  Excel      : {EXCEL_FILE}")
    gs = 'enabled' if GOOGLE_SHEET_ID else 'disabled'
    print(f"  Google Sheets: {gs}")
    print()

    sync()   # run immediately

    schedule.every(INTERVAL_MINUTES).minutes.do(sync)
    while True:
        schedule.run_pending()
        time.sleep(30)
